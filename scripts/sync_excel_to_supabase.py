# -*- coding: utf-8 -*-
"""
sync_excel_to_supabase.py
把 cat_foods_database.xlsx 完整同步到 Supabase cat_foods 表。
策略：先清空，再全部 insert。
"""
import io, sys, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl
from supabase import create_client

SUPABASE_URL = "https://hltnbcqcsmchmfwdcpoc.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImhsdG5iY3Fjc21jaG1md2RjcG9jIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODI1NTAzOTgsImV4cCI6MjA5ODEyNjM5OH0.fsWtKXvdg4dtUVlgvXAQ-IK8T5Eq1QtF06TCYMRhyN0"

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "cat_foods_database.xlsx")

# Supabase CatFood 欄位（來自 lib/supabase.ts）
# Excel 欄位 → Supabase 欄位
COL_MAP = {
    "品牌":           "brand",
    "飼料名稱":       "name",
    "適用階段":       "life_stage",
    "總分":           "score_total",
    "評級":           "score_label",
    "蛋白質% (乾重)": "protein_dm_pct",
    "脂肪% (乾重)":   "fat_dm_pct",
    "碳水% (乾重)":   "carb_dm_pct",
    "水分%":          "moisture_pct",
    "蛋白質% (標示)": "protein_pct",
    "脂肪% (標示)":   "fat_pct",
    "纖維% (標示)":   "fiber_pct",
    "灰分% (標示)":   "ash_pct",
    "含穀物":         "has_grain",
    "AAFCO 認證":     "is_aafco_certified",
    "來源網址":       "source_url",
    "UUID":           "id",
    # 內容評價 → ai_summary（特殊處理，見下方）
}

BOOL_COLS = {"has_grain", "is_aafco_certified"}

# 固定預設值（Supabase not-null 欄位）
DEFAULTS = {
    "food_type":      "dry",
    "has_allergen":   False,
    "has_4d_meat":    False,
}


def to_bool(val):
    if isinstance(val, bool): return val
    if isinstance(val, str): return val.strip() in ("是", "Y", "y", "true", "True", "1")
    return False


def parse_ai_summary(text: str) -> dict | None:
    """把三行評價文字轉成 {good, warning, bad} JSON 格式"""
    if not text:
        return None
    lines = [l.strip() for l in text.strip().split("\n") if l.strip()]
    result = {}
    if len(lines) >= 1:
        result["good"] = lines[0]
    if len(lines) >= 2:
        result["warning"] = lines[1]
    if len(lines) >= 3:
        result["bad"] = lines[2]
    return result if result else None


def main():
    # 讀 Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    rows_data = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(vals):
            continue
        row_dict = {headers[i]: vals[i] for i in range(len(headers)) if headers[i]}
        rows_data.append(row_dict)

    print(f"Excel 讀取：{len(rows_data)} 筆")

    # 過濾 J/K/L（蛋白質/脂肪/纖維 標示值）任一為空的列
    before = len(rows_data)
    rows_data = [r for r in rows_data if all([
        r.get("蛋白質% (標示)"), r.get("脂肪% (標示)"), r.get("纖維% (標示)")
    ])]
    print(f"排除 J/K/L 任一為空：{before - len(rows_data)} 筆，佈署 {len(rows_data)} 筆")

    # 轉換為 Supabase 格式
    sb_rows = []
    for row in rows_data:
        sb = dict(DEFAULTS)

        # 一般欄位對應
        for xl_col, sb_col in COL_MAP.items():
            val = row.get(xl_col)
            if sb_col in BOOL_COLS:
                val = to_bool(val)
            elif isinstance(val, str):
                val = val.strip() or None
            sb[sb_col] = val

        # score_total 確保是 int
        if sb.get("score_total") is not None:
            try:
                sb["score_total"] = int(sb["score_total"])
            except (ValueError, TypeError):
                sb["score_total"] = None

        # 內容評價 → ai_summary (JSON)
        review_text = row.get("內容評價")
        sb["ai_summary"] = parse_ai_summary(review_text) if review_text else None

        # 確保有 id
        if not sb.get("id"):
            import uuid
            sb["id"] = str(uuid.uuid4())

        sb_rows.append(sb)

    # 連線 Supabase
    client = create_client(SUPABASE_URL, SUPABASE_KEY)

    # 清空現有資料
    print("清空 Supabase cat_foods 表...")
    try:
        client.table("cat_foods").delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
        print("  ✅ 清空完成")
    except Exception as e:
        print(f"  ❌ 清空失敗：{e}")
        sys.exit(1)

    # 批次 insert（每批 50 筆）
    print(f"\n開始上傳 {len(sb_rows)} 筆...")
    BATCH = 50
    success = 0
    for i in range(0, len(sb_rows), BATCH):
        batch = sb_rows[i:i+BATCH]
        try:
            client.table("cat_foods").insert(batch).execute()
            success += len(batch)
            print(f"  ✅ [{i+1}~{i+len(batch)}] 上傳成功")
        except Exception as e:
            print(f"  ❌ [{i+1}~{i+len(batch)}] 批次失敗，逐筆重試... 錯誤：{e}")
            for row in batch:
                try:
                    client.table("cat_foods").insert(row).execute()
                    success += 1
                except Exception as e2:
                    print(f"    ❌ {row.get('name','?')} 失敗：{e2}")

    print(f"\n完成：成功 {success}/{len(sb_rows)} 筆上傳到 Supabase")


if __name__ == "__main__":
    main()
