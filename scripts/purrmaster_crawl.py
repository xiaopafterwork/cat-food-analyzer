# -*- coding: utf-8 -*-
"""
purrmaster_crawl.py
爬取 https://purrmaster.com/brand/completedry 全部貓乾糧
- 品牌列表 / 品牌頁：requests（靜態）
- 產品詳情頁：Playwright（JS 渲染）
輸出到 data/purrmaster_database.xlsx（獨立新檔）
"""
import io, json, os, re, sys, time, uuid
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

BASE     = "https://purrmaster.com"
LIST_URL = f"{BASE}/brand/completedry"
DELAY    = 1.5       # 秒，靜態頁面
PW_WAIT  = 3000      # ms，Playwright 等待
HEADERS  = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
}

SCRIPT_DIR = os.path.dirname(__file__)
OUT_EXCEL  = os.path.join(SCRIPT_DIR, "..", "data", "purrmaster_database.xlsx")

GRAIN_KW = ["玉米", "小麥", "大麥", "燕麥", "白米", "糙米", "高粱",
            "corn", "wheat", "barley", "oat", "rice", "sorghum"]

LIFE_STAGE_MAP = {
    "幼貓": ["幼貓", "kitten", "幼齡", "離乳", "成長"],
    "熟齡貓": ["熟齡", "老貓", "senior", "7+", "高齡", "11+"],
    "全齡": ["全齡", "all life", "all age", "各年齡", "成幼"],
}

COLUMNS = [
    ("brand",            "品牌"),
    ("name",             "飼料名稱"),
    ("life_stage",       "適用階段"),
    ("country",          "產地"),
    ("protein_pct",      "蛋白質% (標示)"),
    ("fat_pct",          "脂肪% (標示)"),
    ("fiber_pct",        "纖維% (標示)"),
    ("moisture_pct",     "水分%"),
    ("ash_pct",          "灰分% (標示)"),
    ("carb_pct",         "碳水% (標示)"),
    ("protein_dm_pct",   "蛋白質% (乾重)"),
    ("fat_dm_pct",       "脂肪% (乾重)"),
    ("carb_dm_pct",      "碳水% (乾重)"),
    ("ash_dm_pct",       "灰分% (乾重)"),
    ("fiber_dm_pct",     "纖維% (乾重)"),
    ("has_grain",        "含穀物"),
    ("has_ingredients",  "有成分表"),
    ("ingredients_raw",  "成分原料"),
    ("source_url",       "來源網址"),
    ("id",               "UUID"),
]

COL_WIDTHS = {
    "品牌": 14, "飼料名稱": 35, "適用階段": 8, "產地": 8,
    "蛋白質% (標示)": 11, "脂肪% (標示)": 11, "纖維% (標示)": 11,
    "水分%": 8, "灰分% (標示)": 11, "碳水% (標示)": 11,
    "蛋白質% (乾重)": 11, "脂肪% (乾重)": 11, "碳水% (乾重)": 11,
    "灰分% (乾重)": 11, "纖維% (乾重)": 11,
    "含穀物": 8, "有成分表": 8, "成分原料": 50,
    "來源網址": 45, "UUID": 36,
}


# ── HTTP（靜態頁面）────────────────────────────────────────────
def get_soup(url: str, retries=3):
    for i in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, timeout=20)
            r.raise_for_status()
            r.encoding = "utf-8"
            return BeautifulSoup(r.text, "html.parser")
        except Exception as e:
            if i < retries - 1:
                time.sleep(2)
            else:
                print(f"  ❌ 無法取得 {url[:60]}：{e}")
    return None


# ── 工具函式 ──────────────────────────────────────────────────
def parse_num(text) -> float | None:
    if not text:
        return None
    s = re.sub(r"[%％\s]", "", str(text))
    m = re.search(r"(\d+\.?\d*)", s)
    return float(m.group(1)) if m else None


def guess_life_stage(text: str) -> str:
    t = text.lower()
    for stage, kws in LIFE_STAGE_MAP.items():
        if any(k.lower() in t for k in kws):
            return stage
    return "成貓"


# ── Step 1：取所有品牌連結（requests）────────────────────────
def get_brand_links() -> list[tuple[str, str]]:
    soup = get_soup(LIST_URL)
    if not soup:
        return []
    brands, seen = [], set()
    for a in soup.find_all("a", href=re.compile(r"/brand/completedry/[0-9a-f-]{36}")):
        href = a.get("href", "")
        name = a.get_text(strip=True)
        if href and name:
            url = BASE + href if href.startswith("/") else href
            if url not in seen:
                seen.add(url)
                brands.append((name, url))
    print(f"找到 {len(brands)} 個品牌")
    return brands


# ── Step 2：取品牌下的產品連結（requests）────────────────────
def get_product_links(brand_url: str) -> list[str]:
    soup = get_soup(brand_url)
    if not soup:
        return []
    links, seen = [], set()
    for a in soup.find_all("a", href=re.compile(r"/product/completedry/[0-9a-f-]{36}")):
        href = a.get("href", "")
        url  = BASE + href if href.startswith("/") else href
        if url not in seen:
            seen.add(url)
            links.append(url)
    return links


# ── Step 3：爬產品詳情頁（Playwright）────────────────────────
def scrape_product(page, brand_name: str, url: str) -> dict | None:
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(PW_WAIT)
    except PlaywrightTimeout:
        print(f"  ⏱ 超時：{url[:60]}")
        return None

    # ── 取營養表格 ──
    table_data = page.evaluate("""() => {
        const table = document.querySelector('table');
        if (!table) return null;
        const rows = [...table.querySelectorAll('tr')];
        if (rows.length < 2) return null;

        const headers = [...rows[0].querySelectorAll('th, td')]
            .map(c => c.textContent.replace(/\\s+/g,'').trim());

        const result = { headers };
        rows.slice(1).forEach(row => {
            const cells = [...row.querySelectorAll('th, td')];
            if (!cells.length) return;
            const label = cells[0].textContent.replace(/\\s+/g,'').trim();
            result[label] = cells.slice(1).map(c =>
                c.textContent.replace(/[%％\\s]/g,'').trim()
            );
        });
        return result;
    }""")

    # ── 取頁面文字（品名、產地、成分）──
    body_text = page.inner_text("body")

    # 品名：在「撰寫評論」之後、成分原料之前的非空行
    name = ""
    name_m = re.search(r"撰寫評論\s*\n+([^\n]{3,60})\n", body_text)
    if name_m:
        name = name_m.group(1).strip()
    if not name:
        # fallback: h1
        h1 = page.query_selector("h1")
        name = h1.inner_text().strip() if h1 else ""
    if not name:
        # fallback2: 頁面 title（去掉網站名稱部分）
        title = page.title()
        name = re.sub(r"\s*[|\-].*$", "", title).strip()
    if not name:
        return None

    # 解析表格值
    def col_val(row_key: str, col_name: str) -> float | None:
        if not table_data or row_key not in table_data:
            return None
        headers = table_data.get("headers", [])
        try:
            idx = headers.index(col_name) - 1  # -1 因為 headers 含 row label
            vals = table_data[row_key]
            return parse_num(vals[idx]) if idx < len(vals) else None
        except (ValueError, IndexError):
            return None

    # 分析成分（標示值）
    protein_pct  = col_val("分析成分", "蛋白質")
    fat_pct      = col_val("分析成分", "脂肪")
    fiber_pct    = col_val("分析成分", "纖維")
    moisture_pct = col_val("分析成分", "水份")
    ash_pct      = col_val("分析成分", "灰質")
    carb_pct     = col_val("分析成分", "碳水")

    # 乾物比（直接取站上計算好的值）
    protein_dm = col_val("乾物比(DM)", "蛋白質")
    fat_dm     = col_val("乾物比(DM)", "脂肪")
    fiber_dm   = col_val("乾物比(DM)", "纖維")
    ash_dm     = col_val("乾物比(DM)", "灰質")
    carb_dm    = col_val("乾物比(DM)", "碳水")

    # 至少要有蛋白質才算有效
    if not protein_pct and not protein_dm:
        return None

    # 產地
    country_m = re.search(r"產地\s*[：:]\s*([^\n\t]{2,10})", body_text)
    country = country_m.group(1).strip() if country_m else None

    # 成分原料（在表格之前的長文字）
    # purrmaster 成分列在品名後、「其他」資訊前
    ing_m = re.search(
        r"(?:新鮮|脫水|雞肉|魚肉|肉粉)(.{100,2000}?)(?:其他\s*:|產地\s*:|技術性添加)",
        body_text, re.DOTALL
    )
    if not ing_m:
        # fallback：找第一行很長的文字（成分表通常很長）
        ing_m2 = re.search(r"((?:新鮮|脫水)[^\n]{100,})", body_text)
        ingredients_raw = ing_m2.group(0).strip()[:600] if ing_m2 else None
    else:
        ingredients_raw = ("新鮮" + re.sub(r"\s+", " ", ing_m.group(0)))[:600]

    has_grain = any(k in (ingredients_raw or "").lower() for k in GRAIN_KW)

    return {
        "id":              str(uuid.uuid4()),
        "brand":           brand_name,
        "name":            name,
        "life_stage":      guess_life_stage(name + " " + body_text[:300]),
        "country":         country,
        "source_url":      url,
        "protein_pct":     protein_pct,
        "fat_pct":         fat_pct,
        "fiber_pct":       fiber_pct,
        "moisture_pct":    moisture_pct,
        "ash_pct":         ash_pct,
        "carb_pct":        carb_pct,
        "protein_dm_pct":  protein_dm,
        "fat_dm_pct":      fat_dm,
        "carb_dm_pct":     carb_dm,
        "ash_dm_pct":      ash_dm,
        "fiber_dm_pct":    fiber_dm,
        "has_grain":       has_grain,
        "has_ingredients": bool(ingredients_raw),
        "ingredients_raw": ingredients_raw,
    }


# ── 輸出 Excel ────────────────────────────────────────────────
def save_excel(foods: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purrmaster 貓飼料"

    hfill = PatternFill("solid", fgColor="1B3A2D")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    for ci, (_, label) in enumerate(COLUMNS, 1):
        c = ws.cell(1, ci, label)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    import re as _re
    _illegal = _re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

    def _clean(v):
        if isinstance(v, str):
            return _illegal.sub('', v)
        return v

    for ri, food in enumerate(foods, 2):
        for ci, (key, _) in enumerate(COLUMNS, 1):
            val = food.get(key)
            if isinstance(val, bool):
                val = "是" if val else "否"
            ws.cell(ri, ci, _clean(val)).alignment = Alignment(
                vertical="center", wrap_text=(key == "ingredients_raw")
            )

    for ci, (_, label) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(label, 12)

    ws.freeze_panes = "A2"
    wb.save(OUT_EXCEL)
    print(f"✅ Excel 已儲存：{OUT_EXCEL}（{len(foods)} 筆）")


# ── 主流程 ────────────────────────────────────────────────────
def main():
    print("=" * 55)
    print("  Purrmaster 全站爬蟲（乾糧）")
    print("=" * 55)

    # Step 1：品牌列表（requests）
    print(f"\n── Step 1：取品牌列表 ──────────────────────────────────")
    brands = get_brand_links()
    if not brands:
        print("❌ 無法取得品牌列表，結束")
        return

    # Step 2 + 3：逐品牌爬產品（Playwright 單一 browser 共用）
    print(f"\n── Step 2-3：爬各品牌產品 ──────────────────────────────")
    all_foods    = []
    no_data      = []
    product_urls = set()

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx     = browser.new_context(user_agent=HEADERS["User-Agent"])
        page    = ctx.new_page()

        for bi, (brand_name, brand_url) in enumerate(brands, 1):
            print(f"\n[品牌 {bi}/{len(brands)}] {brand_name}")
            prod_urls = get_product_links(brand_url)
            print(f"  找到 {len(prod_urls)} 個產品")
            time.sleep(DELAY)

            for pi, purl in enumerate(prod_urls, 1):
                if purl in product_urls:
                    continue
                product_urls.add(purl)

                food = scrape_product(page, brand_name, purl)
                if food:
                    all_foods.append(food)
                    print(f"  [{pi}] ✅ {food['name'][:35]} | 蛋白={food.get('protein_pct')}% 水分={food.get('moisture_pct')}%")
                else:
                    no_data.append(purl)
                    print(f"  [{pi}] ⚠️  無成分資料")
                time.sleep(DELAY)

        browser.close()

    # 輸出
    print(f"\n── Step 4：輸出 ─────────────────────────────────────────")
    print(f"成功解析：{len(all_foods)} 筆")
    print(f"無成分資料（略過）：{len(no_data)} 筆")

    if all_foods:
        raw_dir  = os.path.join(SCRIPT_DIR, "..", "data", "raw")
        os.makedirs(raw_dir, exist_ok=True)
        raw_path = os.path.join(raw_dir, f"purrmaster_{time.strftime('%Y%m%d_%H%M')}.json")
        with open(raw_path, "w", encoding="utf-8") as f:
            json.dump(all_foods, f, ensure_ascii=False, indent=2)
        print(f"JSON 備份：{raw_path}")
        save_excel(all_foods)

    print(f"\n完成！{len(all_foods)} 筆資料存入 purrmaster_database.xlsx")


if __name__ == "__main__":
    main()
