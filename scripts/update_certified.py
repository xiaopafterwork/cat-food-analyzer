"""
根據網路查詢結果，標記各品牌 AAFCO / FEDIAF 認證狀況
更新 purrmaster_database.xlsx，新增 is_certified 欄位
"""
import json
import openpyxl
from pathlib import Path

# ── 已確認有 AAFCO 或 FEDIAF 認證的品牌（英文前綴，不分大小寫） ──
CERTIFIED_PREFIXES = [
    # 國際大廠
    'royal canin', "hill's", 'hills', 'proplan', 'pro plan',
    # 美國品牌
    'wellness', 'nulo', 'blackwood', 'instinct', 'natural balance',
    'halo', 'solid gold', 'open farm', 'merrick', 'earthborn',
    'chicken soup', 'canidae', 'blue buffalo', 'avoderm', 'kirkland',
    'diamond naturals', 'taste of the wild', 'annamaet', "grandma mae's",
    'evolve', 'nutri source', 'organix', 'nutro', 'orijen', 'acana',
    'best breed', 'bravo', 'health extension', "dr.tim", 'verus',
    'harlow blend', 'holistic select', 'now!', 'now fresh',
    # 加拿大品牌
    'boreal', 'nutrience', 'go!', 'pronature', '1st choice', 'carna4',
    'firstmate', 'north paw', 'loveabowl', 'nutram', 'tundra',
    # 紐西蘭品牌
    'ziwipeak', 'k9 natural', 'kiwi kitchens', 'addiction', 'zeal',
    # 澳洲品牌
    'vetalogica', 'ivory coat', 'advance', 'kaniva', 'regal',
    # 台灣品牌（確認有 AAFCO feeding trial）
    'blue bay', 'real power',
    # 亞洲品牌
    'anf',
    # 歐洲品牌（FEDIAF，計入認證）
    'n&d farmina', 'farmina', 'monge', 'alleva', 'renske',
    'ownat', 'natural greatness', 'aatu', 'applaws', 'equilibrio',
    'krave',
    # 其他確認品牌
    'wild islands',   # ADDICTION 旗下系列
]

def is_certified(brand: str) -> bool:
    b = brand.lower().strip()
    for prefix in CERTIFIED_PREFIXES:
        if b.startswith(prefix) or prefix in b:
            return True
    return False

# ── 載入 JSON ──
json_path = Path('data/raw/purrmaster_20260629_1518.json')
with open(json_path, encoding='utf-8') as f:
    foods = json.load(f)

# ── 統計 ──
certified_brands = set()
uncertified_brands = set()
for food in foods:
    brand = food.get('brand', '')
    if is_certified(brand):
        certified_brands.add(brand)
    else:
        uncertified_brands.add(brand)

print(f'Certified: {len(certified_brands)}')
print(f'No cert: {len(uncertified_brands)}')
print('\n--- Certified ---')
for b in sorted(certified_brands):
    cnt = sum(1 for f in foods if f.get('brand') == b)
    print(f'  OK [{cnt:3d}] {b[:40]}')

print('\n--- No cert ---')
for b in sorted(uncertified_brands):
    cnt = sum(1 for f in foods if f.get('brand') == b)
    print(f'  ?  [{cnt:3d}] {b[:40]}')

# ── 更新 Excel ──
xl_path = Path('data/purrmaster_database.xlsx')
wb = openpyxl.load_workbook(xl_path)
ws = wb.active

# 找現有標題列
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print(f'\nColumns: {len(headers)}')

# 新增 is_certified 欄
if 'is_certified' not in headers:
    new_col = ws.max_column + 1
    ws.cell(1, new_col, 'is_certified')
else:
    new_col = headers.index('is_certified') + 1

# brand 固定在第 1 欄（確認自資料）
brand_col = 1
print(f'brand col: {brand_col}, is_certified col: {new_col}')

# 逐列更新
cert_count = 0
for row in range(2, ws.max_row + 1):
    brand = ws.cell(row, brand_col).value or ''
    val = is_certified(brand)
    ws.cell(row, new_col, val)
    if val:
        cert_count += 1

print(f'\nCertified rows: {cert_count} / {ws.max_row - 1}')

wb.save(xl_path)
print(f'Saved: {xl_path}')
