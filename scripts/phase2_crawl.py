# -*- coding: utf-8 -*-
"""
phase2_crawl.py — 讀批准 Excel，爬兩個來源，官網數據優先

流程：
  1. 讀取 data/review/ 最新的待審核_*.xlsx
  2. 找出批准欄 = Y 的行（爬蟲清單 + 手動新增都讀）
  3. 每筆爬：① 官網URL  ② PChome 第二來源URL
  4. 合併：官網有的欄位用官網，官網沒有才用第二來源補
  5. 輸出 JSON 到 data/raw/，差異明顯的欄位標記「待確認」
  6. 更新 Excel：在「狀態」欄寫入結果

用法：
  python phase2_crawl.py              # 自動找最新 Excel
  python phase2_crawl.py --file 待審核_20260628.xlsx
"""

import io, json, os, sys, argparse, glob, time, re
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("請先安裝：pip install openpyxl"); sys.exit(1)

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    print("請先安裝：pip install requests beautifulsoup4"); sys.exit(1)

SCRIPT_DIR = os.path.dirname(__file__)
REVIEW_DIR = os.path.join(SCRIPT_DIR, "data/review")
RAW_DIR    = os.path.join(SCRIPT_DIR, "data/raw")
os.makedirs(RAW_DIR, exist_ok=True)

HEADERS_ROW_MARKER = "品牌"   # 用來識別表頭列
APPROVE_COL  = 5              # E 欄：批准
NOTE_COL     = 6              # F 欄：備註（寫入爬蟲結果）
STATUS_FILL_OK   = PatternFill("solid", fgColor="C8E6C9")
STATUS_FILL_WARN = PatternFill("solid", fgColor="FFF9C4")
STATUS_FILL_ERR  = PatternFill("solid", fgColor="FFCCBC")

# ── 欄位差異容忍值（超過才標記「待確認」）────────────────────
TOLERANCE = {
    "protein_pct": 3.0,
    "fat_pct":     3.0,
    "fiber_pct":   2.0,
    "moisture_pct":3.0,
    "ash_pct":     2.0,
}


# ── 通用爬蟲（抓營養成分表）────────────────────────────────
HEADERS_HTTP = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}

def get_soup(url: str):
    try:
        r = requests.get(url, headers=HEADERS_HTTP, timeout=15)
        r.raise_for_status()
        return BeautifulSoup(r.text, "html.parser")
    except Exception as e:
        print(f"    ⚠ 無法取得 {url[:60]}…: {e}")
        return None

NUM_RE = re.compile(r"(\d+\.?\d*)")

def extract_pct(text: str):
    m = NUM_RE.search(str(text))
    return float(m.group(1)) if m else None

NUTRITION_KEYWORDS = {
    "protein_pct":  ["粗蛋白", "crude protein", "蛋白質"],
    "fat_pct":      ["粗脂肪", "crude fat", "脂肪"],
    "fiber_pct":    ["粗纖維", "crude fiber", "纖維"],
    "moisture_pct": ["水分", "moisture"],
    "ash_pct":      ["灰分", "crude ash", "ash"],
}

def parse_nutrition(soup) -> dict:
    """從任意頁面掃描含保證分析值的表格或段落。"""
    result = {k: None for k in NUTRITION_KEYWORDS}
    if soup is None:
        return result

    text = soup.get_text(" ")
    lines = text.splitlines()

    for line in lines:
        line_lower = line.lower()
        for field, keywords in NUTRITION_KEYWORDS.items():
            if result[field] is not None:
                continue
            for kw in keywords:
                if kw in line_lower:
                    val = extract_pct(line)
                    if val is not None:
                        result[field] = val
                    break
    return result

def parse_ingredients(soup) -> str | None:
    if soup is None:
        return None
    text = soup.get_text(" ")
    m = re.search(r"(成分|原料|ingredients)[：:＊*]?\s*(.{20,500})", text, re.IGNORECASE)
    return m.group(2).strip()[:500] if m else None


# ── 合併兩個來源（官網優先）────────────────────────────────
def merge(official: dict, second: dict) -> tuple[dict, list[str]]:
    """回傳合併結果 + 差異警告清單。"""
    merged   = dict(official)
    warnings = []

    for field, tol in TOLERANCE.items():
        o_val = official.get(field)
        s_val = second.get(field)

        if o_val is None and s_val is not None:
            merged[field] = s_val   # 官網缺漏，用第二來源補
        elif o_val is not None and s_val is not None:
            diff = abs(o_val - s_val)
            if diff > tol:
                warnings.append(f"{field}: 官網={o_val} vs 第二來源={s_val}（差 {diff:.1f}%）")

    # 成分以官網為主，沒有才補
    if not merged.get("ingredients_raw") and second.get("ingredients_raw"):
        merged["ingredients_raw"] = second["ingredients_raw"]

    return merged, warnings


# ── Excel 讀取 ──────────────────────────────────────────────
def find_latest_excel() -> str | None:
    files = sorted(glob.glob(os.path.join(REVIEW_DIR, "待審核_*.xlsx")), reverse=True)
    return files[0] if files else None

def read_approved_rows(xlsx_path: str) -> list[dict]:
    wb  = openpyxl.load_workbook(xlsx_path)
    ws  = wb.active
    approved = []
    in_data  = False

    for row in ws.iter_rows():
        first = str(row[0].value or "").strip()

        # 識別表頭列（可能出現兩次：爬蟲清單 + 手動新增各一）
        if first == HEADERS_ROW_MARKER:
            in_data = True
            continue

        if not in_data:
            continue

        # 資料列：批准欄 = Y
        approve_val = str(row[APPROVE_COL - 1].value or "").strip().upper()
        if approve_val != "Y":
            continue

        brand = str(row[0].value or "").strip()
        name  = str(row[1].value or "").strip()
        if not brand or not name:
            continue

        approved.append({
            "brand":        brand,
            "name":         name,
            "official_url": str(row[2].value or "").strip(),
            "second_url":   str(row[3].value or "").strip(),
            "_row":         row[0].row,   # 記錄行號，之後回寫
        })

    return approved, wb, ws


# ── 主流程 ───────────────────────────────────────────────────
def crawl_one(item: dict) -> tuple[dict, list[str], str]:
    brand = item["brand"]
    name  = item["name"]
    o_url = item["official_url"]
    s_url = item["second_url"]

    print(f"\n  [{brand}] {name}")

    # 官網
    official_raw = {}
    if o_url:
        print(f"    官網: {o_url[:70]}")
        soup = get_soup(o_url)
        official_raw = parse_nutrition(soup)
        official_raw["ingredients_raw"] = parse_ingredients(soup)
        official_raw["source_url"] = o_url
        time.sleep(1)
    else:
        print("    ⚠ 未填官網URL，跳過官網爬蟲")

    # 第二來源
    second_raw = {}
    if s_url:
        print(f"    第二來源: {s_url[:70]}")
        soup2 = get_soup(s_url)
        second_raw = parse_nutrition(soup2)
        second_raw["ingredients_raw"] = parse_ingredients(soup2)
        time.sleep(1)

    merged, warnings = merge(official_raw, second_raw)
    merged.update({
        "brand":      brand,
        "name":       name,
        "second_url": s_url,
    })

    status = "ok" if not warnings else "warn"
    if not o_url and not s_url:
        status = "error"

    return merged, warnings, status


def save_json(data: dict) -> str:
    safe_name = re.sub(r"[^\w一-鿿]", "_", f"{data['brand']}_{data['name']}")[:50]
    ts = datetime.now().strftime("%H%M%S")
    fname = f"{safe_name}_{ts}.json"
    path  = os.path.join(RAW_DIR, fname)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path


def update_excel_row(ws, row_num: int, warnings: list[str], status: str):
    fill = STATUS_FILL_OK if status == "ok" else (STATUS_FILL_WARN if status == "warn" else STATUS_FILL_ERR)
    note = "✅ 完成" if not warnings else "⚠ 差異：" + " | ".join(warnings)
    c = ws.cell(row=row_num, column=NOTE_COL, value=note)
    c.fill = fill
    c.font = Font(size=9)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", help="指定 Excel 檔名（放在 data/review/）")
    args = parser.parse_args()

    if args.file:
        xlsx_path = os.path.join(REVIEW_DIR, args.file)
    else:
        xlsx_path = find_latest_excel()

    if not xlsx_path or not os.path.exists(xlsx_path):
        print("找不到 Excel，請先執行 python phase1_export.py")
        sys.exit(1)

    print(f"讀取：{xlsx_path}")
    approved, wb, ws = read_approved_rows(xlsx_path)

    if not approved:
        print("沒有批准（Y）的列，結束。")
        sys.exit(0)

    print(f"\n共 {len(approved)} 筆待爬蟲\n{'─'*50}")
    total_ok = total_warn = total_err = 0

    for item in approved:
        data, warnings, status = crawl_one(item)
        json_path = save_json(data)

        if status == "ok":     total_ok   += 1
        elif status == "warn": total_warn += 1
        else:                  total_err  += 1

        update_excel_row(ws, item["_row"], warnings, status)

        if warnings:
            print(f"    ⚠ 差異（官網數據優先使用）：")
            for w in warnings: print(f"      {w}")
        print(f"    → 已儲存 {os.path.basename(json_path)}")

    # 回寫 Excel
    wb.save(xlsx_path)
    print(f"\n{'─'*50}")
    print(f"完成：✅ {total_ok} 筆  ⚠ {total_warn} 筆差異  ❌ {total_err} 筆失敗")
    print(f"Excel 已更新：{xlsx_path}")
    print(f"JSON 輸出至：{RAW_DIR}")
    print("\n下一步：執行 python run_pipeline.py 完成清洗→評分→上傳")


if __name__ == "__main__":
    main()
