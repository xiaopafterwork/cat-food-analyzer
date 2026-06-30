# -*- coding: utf-8 -*-
"""
petpark_fresh_crawl.py
全新爬取 shop.petpark.com.tw 所有貓乾糧，從零建立 Excel。
修正：過濾 CSS 垃圾字串、品名清洗、統一欄位格式。
"""
import io, json, os, re, sys, time, uuid
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from playwright.sync_api import sync_playwright, TimeoutError as PwTimeout
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(__file__)
EXCEL_PATH = os.path.join(SCRIPT_DIR, "..", "data", "cat_foods_database.xlsx")
RAW_DIR    = os.path.join(SCRIPT_DIR, "data", "raw")
os.makedirs(RAW_DIR, exist_ok=True)

BASE_URL = "https://shop.petpark.com.tw/dryfood/hot-cat-dryfood"
DELAY    = 2.0

# ── 欄位定義 ─────────────────────────────────────────────────
COLUMNS = [
    ("brand",               "品牌"),
    ("name",                "飼料名稱"),
    ("life_stage",          "適用階段"),
    ("score_total",         "總分"),
    ("score_label",         "評級"),
    ("protein_dm_pct",      "蛋白質% (乾重)"),
    ("fat_dm_pct",          "脂肪% (乾重)"),
    ("carb_dm_pct",         "碳水% (乾重)"),
    ("moisture_pct",        "水分%"),
    ("protein_pct",         "蛋白質% (標示)"),
    ("fat_pct",             "脂肪% (標示)"),
    ("fiber_pct",           "纖維% (標示)"),
    ("ash_pct",             "灰分% (標示)"),
    ("has_grain",           "含穀物"),
    ("is_aafco_certified",  "AAFCO 認證"),
    ("has_ingredient_list", "有成分表"),
    ("has_ash_listed",      "有標示灰分"),
    ("score_protein",       "蛋白質分"),
    ("score_carb",          "碳水分"),
    ("score_fat",           "脂肪分"),
    ("score_transparency",  "透明度分"),
    ("score_grain",         "無穀分"),
    ("score_ash_quality",   "灰分品質分"),
    ("source_url",          "來源網址"),
    ("id",                  "UUID"),
]
COL_WIDTHS = {
    "品牌": 14, "飼料名稱": 35, "適用階段": 8, "總分": 7, "評級": 10,
    "蛋白質% (乾重)": 12, "脂肪% (乾重)": 12, "碳水% (乾重)": 12,
    "水分%": 8,
    "蛋白質% (標示)": 12, "脂肪% (標示)": 12, "纖維% (標示)": 12, "灰分% (標示)": 12,
    "含穀物": 8, "AAFCO 認證": 10, "有成分表": 8, "有標示灰分": 10,
    "蛋白質分": 8, "碳水分": 8, "脂肪分": 8, "透明度分": 8, "無穀分": 8, "灰分品質分": 10,
    "來源網址": 40, "UUID": 36,
}

STAGE_LABEL = {"kitten": "幼貓", "adult": "成貓", "senior": "熟齡貓", "all": "全齡"}
GRAIN_KW    = ["玉米", "小麥", "大麥", "燕麥", "白米", "糙米", "高粱",
               "corn", "wheat", "barley", "oat", "rice", "sorghum"]

LIFE_STAGE_MAP = {
    "kitten": ["幼貓", "kitten", "幼齡", "離乳"],
    "senior": ["熟齡", "老貓", "senior", "7+", "高齡", "11+"],
    "all":    ["全齡", "all life", "all age", "各年齡"],
}

# ── 品名清洗 ─────────────────────────────────────────────────
NOISE_PATTERNS = [
    r"\s*\(貓飼料\)", r"\s*（貓飼料）", r"\s*\(貓糧\)",
    r"\s*\d+\.?\d*\s*(公斤|公克|kg|KG|磅|lb|LB|克|g|G|ml|ML|L|入|包|袋)\b.*",
    r"\s+\d+\s*$",
]

def clean_name(name: str) -> str:
    for pat in NOISE_PATTERNS:
        name = re.sub(pat, "", name, flags=re.IGNORECASE)
    return name.strip()

def is_garbage(name: str) -> bool:
    """過濾 CSS / JS / HTML 垃圾字串"""
    if not name or len(name) < 3:
        return True
    if name.startswith(".") or "{" in name or "}" in name:
        return True
    if re.match(r"^\s*[\.\#]", name):
        return True
    if re.search(r"width\s*:|height\s*:|display\s*:", name):
        return True
    return False

# ── 評分公式 ─────────────────────────────────────────────────
def get_label(t):
    if t >= 80: return "優質主食"
    if t >= 65: return "不錯的選擇"
    if t >= 50: return "可以接受"
    return "需謹慎"

def calc_dm(pct, moisture):
    if pct is None or moisture is None or moisture >= 100: return None
    return round(pct / (1 - moisture / 100), 1)

def calc_carb_dm(p, f, a, fi):
    if any(v is None for v in [p, f, a]): return None
    return round(max(0, 100 - p - f - a - (fi or 0)), 1)

def score_food(food):
    p  = food.get("protein_dm_pct") or 0
    c  = food.get("carb_dm_pct") or 100
    f  = food.get("fat_dm_pct") or 0
    a  = food.get("ash_pct") or 0
    has_grain  = food.get("has_grain", False)
    is_aafco   = food.get("is_aafco_certified", False)
    has_ing    = bool(food.get("has_ingredient_list"))
    has_ash    = food.get("has_ash_listed", False)
    has_fat    = food.get("fat_dm_pct") is not None

    if   p >= 50: pp = 30
    elif p >= 45: pp = 26
    elif p >= 40: pp = 22
    elif p >= 35: pp = 20
    elif p >= 30: pp = 13
    elif p >= 26: pp = 7
    else:         pp = 2

    if   c <= 10: cp = 15
    elif c <= 20: cp = 12
    elif c <= 30: cp = 9
    elif c <= 40: cp = 8
    else:         cp = 2

    if has_fat and f > 0:
        if   20 <= f <= 40: fp = 10
        elif 15 <= f < 20 or 40 < f <= 50: fp = 7
        elif  9 <= f < 15: fp = 4
        else: fp = 1
    else: fp = 0

    tp = (18 if is_aafco else 0) + (7 if has_ing else 0) + (5 if has_ash else 0)
    gp = 0 if has_grain else 8
    if has_ash:
        ap = 7 if a <= 8 else (3 if a <= 10 else 0)
    else: ap = 0

    total = max(0, min(100, pp + cp + fp + tp + gp + ap))
    food.update({
        "score_total": total, "score_label": get_label(total),
        "score_protein": pp, "score_carb": cp, "score_fat": fp,
        "score_transparency": tp, "score_grain": gp, "score_ash_quality": ap,
    })
    return food

# ── 成分解析 ─────────────────────────────────────────────────
def parse_pct(text, keyword):
    m = re.search(rf"{keyword}[^0-9]{{0,20}}?(\d+\.?\d*)\s*[％%]", text)
    return float(m.group(1)) if m else None

def guess_life_stage(text):
    t = text.lower()
    for stage, kws in LIFE_STAGE_MAP.items():
        if any(k.lower() in t for k in kws):
            return stage
    return "adult"

KNOWN_BRANDS = [
    "法國皇家", "Royal Canin", "希爾思", "Hills", "冠能", "PRO PLAN", "Purina",
    "耐吉斯", "Nutrience", "自然平衡", "奇境", "魏大夫", "藍饌", "Blue Buffalo",
    "愛肯拿", "ACANA", "歐睿健", "Orijen", "ORIJEN", "Farmina", "法米納",
    "Wellness", "威樂", "汪喵星球", "御天貓", "原點", "無敵", "怡親",
    "K9 Natural", "K9Natural", "Petlife", "晶饌", "鮮樂嚐", "惜時", "Seeds",
    "Addiction", "自然癮食", "滋益巔峰", "Ziwi", "ZIWI", "Go!", "GO!",
    "Instinct", "本能", "史黛拉", "Stella", "Merrick", "瑪瑞克",
    "Canidae", "卡比", "Solid Gold", "固德", "Nulo", "紐樂",
    "Fromm", "法蘭", "Nutro", "美士", "Eukanuba", "優卡",
    "NOW Fresh", "原鮮", "Iams", "寵愛一生", "Pronature",
    "CHARM", "野性魅力", "ADD", "愛德勝", "Kit Cat", "怪獸部落",
    "Catopia", "樂境", "OpenFarm", "開放農場", "Halo", "曙光",
    "優格", "銀湯匙", "MONGE", "瑪恩吉", "Brit", "布莉特",
    "Carnilove", "卡尼洛夫", "Zignature", "真特",
]

def guess_brand(name):
    for b in KNOWN_BRANDS:
        if b.lower() in name.lower():
            return b
    m = re.match(r"^([^\s（(【\[]+)", name)
    return m.group(1) if m else name[:6]

# ── 爬商品清單 ───────────────────────────────────────────────
def get_all_links(page):
    all_items = []
    p = 1
    while True:
        url = BASE_URL if p == 1 else f"{BASE_URL}?p={p}"
        print(f"  分頁 {p}：{url}")
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=45000)
            page.wait_for_timeout(4000)
        except PwTimeout:
            print(f"  ⚠️  分頁 {p} 逾時，停止")
            break

        # 只取 product-item-name 內的 <a>，避免抓到其他雜項連結
        items = page.evaluate("""() => {
            const results = [];
            // 精準選取商品名稱連結
            const selectors = [
                '.product-item-name a',
                '.product-name a',
                'strong.product-item-name a',
                'a.product-item-link',
            ];
            const seen = new Set();
            for (const sel of selectors) {
                document.querySelectorAll(sel).forEach(a => {
                    const href = a.href || '';
                    const name = (a.textContent || '').trim();
                    if (name && href && href.includes('petpark') && !seen.has(href)) {
                        // 排除 CSS/JS 垃圾
                        if (!name.startsWith('.') && !name.includes('{')) {
                            seen.add(href);
                            results.push({ name, href });
                        }
                    }
                });
            }
            return results;
        }""")

        # 額外過濾：名稱不能是 CSS 或空白
        items = [i for i in items if not is_garbage(i["name"])]

        if not items:
            print(f"  分頁 {p} 無商品，結束")
            break

        all_items.extend(items)
        print(f"    取得 {len(items)} 筆，累計 {len(all_items)} 筆")

        has_next = page.evaluate(f"""() => !!document.querySelector('a[href*="?p={p+1}"]')""")
        if not has_next:
            break
        p += 1
        time.sleep(DELAY)

    return all_items

# ── 爬單一商品頁 ─────────────────────────────────────────────
def scrape_product(page, name, url):
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=40000)
        page.wait_for_timeout(2000)
    except PwTimeout:
        return None

    text = page.inner_text("body")

    protein_pct  = parse_pct(text, "粗蛋白質?")
    fat_pct      = parse_pct(text, "粗脂肪")
    fiber_pct    = parse_pct(text, "粗纖維")
    moisture_pct = parse_pct(text, "水分")
    ash_pct      = parse_pct(text, "粗?灰分")

    if not any([protein_pct, fat_pct, moisture_pct]):
        return None

    ing_m = re.search(
        r"(?:成分|原料)[：:＊*]?\s*(.{20,600}?)(?:\n\n|保證|分析|商品規格|$)",
        text, re.DOTALL
    )
    ingredients_raw = re.sub(r"\s+", " ", ing_m.group(1).strip())[:500] if ing_m else None

    is_aafco  = bool(re.search(r"AAFCO|aafco", text))
    has_grain = any(k in (ingredients_raw or "").lower() for k in GRAIN_KW)

    m_pct = moisture_pct
    p_dm  = calc_dm(protein_pct, m_pct)
    f_dm  = calc_dm(fat_pct, m_pct)
    fi_dm = calc_dm(fiber_pct, m_pct)
    a_dm  = calc_dm(ash_pct, m_pct)
    c_dm  = calc_carb_dm(p_dm, f_dm, a_dm, fi_dm)

    food = {
        "id":               str(uuid.uuid4()),
        "name":             name,
        "brand":            guess_brand(name),
        "life_stage":       guess_life_stage(name + " " + text[:300]),
        "food_type":        "dry",
        "source_url":       url,
        "ingredients_raw":  ingredients_raw,
        "protein_pct":      protein_pct,
        "fat_pct":          fat_pct,
        "fiber_pct":        fiber_pct,
        "moisture_pct":     moisture_pct,
        "ash_pct":          ash_pct,
        "protein_dm_pct":   p_dm,
        "fat_dm_pct":       f_dm,
        "carb_dm_pct":      c_dm,
        "has_grain":        has_grain,
        "is_aafco_certified": is_aafco,
        "has_ingredient_list": bool(ingredients_raw),
        "has_ash_listed":   ash_pct is not None,
    }
    return score_food(food)

# ── 建立新 Excel ─────────────────────────────────────────────
def create_excel(foods):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "貓飼料資料庫"

    hdr_fill = PatternFill("solid", fgColor="2D4A2E")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)

    for ci, (_, label) in enumerate(COLUMNS, 1):
        c = ws.cell(1, ci, label)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    def score_fill(s):
        if s is None: return None
        if s >= 80: return PatternFill("solid", fgColor="C8E6C9")
        if s >= 65: return PatternFill("solid", fgColor="FFF9C4")
        if s >= 50: return PatternFill("solid", fgColor="FFE0B2")
        return PatternFill("solid", fgColor="FFCDD2")

    for ri, food in enumerate(foods, 2):
        s = food.get("score_total")
        food["life_stage"] = STAGE_LABEL.get(food.get("life_stage", "adult"), food.get("life_stage", ""))
        for ci, (key, label) in enumerate(COLUMNS, 1):
            val = food.get(key)
            if isinstance(val, bool):
                val = "是" if val else "否"
            c = ws.cell(ri, ci, val)
            c.alignment = Alignment(vertical="center")
            if label == "總分" and s is not None:
                c.fill = score_fill(s)

    for ci, (_, label) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(label, 12)

    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)
    print(f"✅ Excel 已儲存：{EXCEL_PATH}（{len(foods)} 筆）")

# ── 去重 ─────────────────────────────────────────────────────
def dedup(items):
    seen_urls = set()
    seen_names = set()
    result = []
    for item in items:
        url  = item.get("href", "")
        name = clean_name(item.get("name", ""))
        key  = re.sub(r"\s+", "", name.lower())[:18]
        if url in seen_urls or key in seen_names:
            continue
        seen_urls.add(url)
        seen_names.add(key)
        item["name"] = name
        result.append(item)
    return result

# ── 主流程 ───────────────────────────────────────────────────
def main():
    print("=" * 55)
    print("  Petpark 全新爬蟲")
    print("=" * 55)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="zh-TW",
        )
        list_page   = ctx.new_page()
        detail_page = ctx.new_page()

        print("\n── Step 1：抓商品清單 ──────────────────────────────────")
        all_items = get_all_links(list_page)
        print(f"\n總計找到 {len(all_items)} 個商品（去重前）")

        unique_items = dedup(all_items)
        print(f"去重後：{len(unique_items)} 個\n")

        print("── Step 2：爬成分資料 ──────────────────────────────────")
        results  = []
        no_data  = []

        for i, item in enumerate(unique_items, 1):
            name = item["name"]
            url  = item["href"]
            print(f"[{i}/{len(unique_items)}] {name[:50]}")
            food = scrape_product(detail_page, name, url)
            if food:
                results.append(food)
                print(f"  ✅ 蛋白質={food.get('protein_pct')}%  水分={food.get('moisture_pct')}%  → {food['score_total']}分 [{food['score_label']}]")
            else:
                no_data.append(name)
                print(f"  ⚠️  無成分資料")
            time.sleep(DELAY)

        browser.close()

    print(f"\n── Step 3：輸出 ────────────────────────────────────────")
    print(f"成功解析：{len(results)} 筆")
    print(f"無成分資料（略過）：{len(no_data)} 筆")

    if results:
        raw_path = os.path.join(RAW_DIR, f"petpark_fresh_{time.strftime('%Y%m%d_%H%M')}.json")
        with open(raw_path, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print(f"JSON 備份：{raw_path}")
        create_excel(results)

        # 評分分布
        from collections import Counter
        dist = Counter(r["score_label"] for r in results)
        print("\n評分分布：")
        for lbl in ["優質主食", "不錯的選擇", "可以接受", "需謹慎"]:
            print(f"  {lbl}: {dist.get(lbl, 0)} 筆")

    print(f"\n完成！共 {len(results)} 筆有效資料寫入 Excel。")
    print(f"確認後請執行：python scripts/sync_excel_to_supabase.py")


if __name__ == "__main__":
    main()
