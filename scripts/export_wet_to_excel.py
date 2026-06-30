# -*- coding: utf-8 -*-
"""
export_wet_to_excel.py
把 purrmaster_wet_*.json 輸出成 Excel，表頭格式與 cat_foods_database.xlsx 相同。
新增欄位放最後：食物類型、纖維% (乾重)、每百大卡蛋白質g。
"""
import io, sys, os, json, glob, math
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAW_DIR   = os.path.join(os.path.dirname(__file__), '..', 'data', 'raw')
OUT_DIR   = os.path.join(os.path.dirname(__file__), '..', 'data')
OUT_FILE  = os.path.join(OUT_DIR, 'wet_foods_database.xlsx')


# ── 濕食評分 v4.0（拉高門檻版）──────────────────────────────────────────────

def protein_score(dm):
    if dm is None: return 0
    if dm >= 65: return 50
    if dm >= 60: return 46
    if dm >= 55: return 41
    if dm >= 50: return 35
    if dm >= 45: return 28
    if dm >= 40: return 18
    return 5

def carb_score(dm):
    if dm is None: return 0
    if dm <= 5:  return 30
    if dm <= 10: return 26
    if dm <= 15: return 21
    if dm <= 20: return 18
    if dm <= 30: return 12
    return 5

def quality_score(ash_pct):
    return 5 if ash_pct is not None else 0

def score_label(total):
    if total >= 75: return '優質主食'
    if total >= 60: return '均衡日常'
    if total >= 45: return '基礎配方'
    return '建議搭配'

def protein_per_100kcal(p_dm, f_dm, c_dm):
    if not p_dm: return None
    f_dm = f_dm or 0
    c_dm = c_dm or 0
    total_kcal = p_dm * 3.5 + f_dm * 8.5 + c_dm * 3.5
    if total_kcal == 0: return None
    return round(p_dm / total_kcal * 100, 1)


# ── Excel 樣式 ───────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill('solid', fgColor='1B3A5C')
HEADER_FONT   = Font(bold=True, color='FFFFFF', size=10)
NEW_COL_FILL  = PatternFill('solid', fgColor='EEF3F8')
NEW_COL_FONT  = Font(bold=True, color='1B3A5C', size=10)

SCORE_COLORS = {
    '優質主食': ('E8F9EE', '1A7F37'),
    '均衡日常': ('EDF5F0', '4A7C59'),
    '基礎配方': ('FFF3E0', 'B35C00'),
    '建議搭配': ('FFEAEA', 'C0392B'),
}

thin = Side(style='thin', color='E5E7EB')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def main():
    # 找最新 JSON
    files = sorted(glob.glob(os.path.join(RAW_DIR, 'purrmaster_wet_*.json')))
    if not files:
        print('❌ 找不到 purrmaster_wet_*.json')
        sys.exit(1)
    src = files[-1]
    print(f'來源：{os.path.basename(src)}')

    with open(src, encoding='utf-8') as f:
        raw = json.load(f)

    valid = [r for r in raw if r.get('protein_dm_pct')]
    print(f'有效筆數：{len(valid)}（略過 {len(raw)-len(valid)} 筆無乾物比資料）')

    # ── 表頭（與乾飼料 Excel 相同順序，新增欄位在後）──────────────────────
    HEADERS = [
        '品牌', '飼料名稱', '適用階段',
        '總分', '評級',
        '蛋白質% (乾重)', '脂肪% (乾重)', '碳水% (乾重)', '灰分% (乾重)',
        '水分%',
        '蛋白質% (標示)', '脂肪% (標示)', '纖維% (標示)', '灰分% (標示)',
        '含穀物', 'AAFCO 認證', '有成分表', '有標示灰分',
        '蛋白質分', '碳水分',
        '內容評價',
        '來源網址', 'UUID',
        # ── 濕食新增欄位 ──
        '食物類型', '纖維% (乾重)', '每百大卡蛋白質g',
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '濕食評分資料庫'
    ws.freeze_panes = 'A2'

    # 寫表頭
    NEW_COLS = {'食物類型', '纖維% (乾重)', '每百大卡蛋白質g'}
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, col_idx, h)
        if h in NEW_COLS:
            cell.fill = NEW_COL_FILL
            cell.font = NEW_COL_FONT
        else:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 32

    # 寫資料
    for r_idx, r in enumerate(valid, 2):
        p_dm  = r.get('protein_dm_pct')
        f_dm  = r.get('fat_dm_pct')
        c_dm  = r.get('carb_dm_pct')
        a_dm  = r.get('ash_dm_pct')
        fi_dm = r.get('fiber_dm_pct')
        ash   = r.get('ash_pct') if r.get('ash_pct') else None

        ps    = protein_score(p_dm)
        cs    = carb_score(c_dm)
        qs    = quality_score(ash)
        total = ps + cs + qs
        label = score_label(total)
        p100  = protein_per_100kcal(p_dm, f_dm, c_dm)

        row_data = [
            r.get('brand', '').strip(),
            r.get('name', '').strip(),
            r.get('life_stage') or 'all',
            total,
            label,
            round(p_dm, 1) if p_dm else None,
            round(f_dm, 1) if f_dm else None,
            round(c_dm, 1) if c_dm else None,
            round(a_dm, 1) if a_dm else None,
            r.get('moisture_pct'),
            r.get('protein_pct'),
            r.get('fat_pct'),
            r.get('fiber_pct'),
            ash,
            '是' if r.get('has_grain') else '否',
            '否',  # AAFCO — 濕食暫無資料
            '是' if r.get('has_ingredients') else '否',
            '是' if ash else '否',
            ps,
            cs,
            None,  # 內容評價
            r.get('source_url'),
            r['id'],
            # 新增欄位
            'wet',
            round(fi_dm, 1) if fi_dm else None,
            p100,
        ]

        score_bg, score_fg = SCORE_COLORS.get(label, ('F3F4F6', '6B7280'))

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(r_idx, col_idx, val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center')

            h = HEADERS[col_idx - 1]
            # 評級欄上色
            if h == '評級':
                cell.fill = PatternFill('solid', fgColor=score_bg)
                cell.font = Font(color=score_fg, bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # 總分欄上色
            elif h == '總分':
                cell.fill = PatternFill('solid', fgColor=score_bg)
                cell.font = Font(color=score_fg, bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # 新增欄位淡藍底
            elif h in NEW_COLS:
                cell.fill = PatternFill('solid', fgColor='F0F5FA')

    # 欄寬
    COL_WIDTHS = {
        '品牌': 18, '飼料名稱': 40, '適用階段': 10,
        '總分': 8, '評級': 10,
        '蛋白質% (乾重)': 14, '脂肪% (乾重)': 12, '碳水% (乾重)': 12, '灰分% (乾重)': 12,
        '水分%': 10,
        '蛋白質% (標示)': 14, '脂肪% (標示)': 12, '纖維% (標示)': 12, '灰分% (標示)': 12,
        '含穀物': 10, 'AAFCO 認證': 12, '有成分表': 10, '有標示灰分': 12,
        '蛋白質分': 10, '碳水分': 10,
        '內容評價': 40, '來源網址': 50, 'UUID': 38,
        '食物類型': 10, '纖維% (乾重)': 12, '每百大卡蛋白質g': 16,
    }
    for col_idx, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(h, 12)

    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}1'

    wb.save(OUT_FILE)
    print(f'\n✅ 已輸出：{OUT_FILE}')
    print(f'   筆數：{len(valid)} 筆，{len(HEADERS)} 欄')

    # 分布摘要
    from collections import Counter
    dist = Counter(score_label(
        protein_score(r.get('protein_dm_pct')) +
        carb_score(r.get('carb_dm_pct')) +
        quality_score(r.get('ash_pct') if r.get('ash_pct') else None)
    ) for r in valid)
    print()
    print('分數分布：')
    for lbl in ['優質主食', '均衡日常', '基礎配方', '建議搭配']:
        print(f'  {lbl}: {dist[lbl]} 筆')


if __name__ == '__main__':
    main()
