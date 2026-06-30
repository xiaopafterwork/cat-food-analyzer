# -*- coding: utf-8 -*-
"""
petpark_full_crawl.py
1. 爬 shop.petpark.com.tw 所有貓乾糧商品名稱
2. 對比現有 Excel 資料庫，只抓沒有的新品
3. 解析成分 → 計算乾物比 → 評分
4. 輸出到 data/cat_foods_database.xlsx（新增列在最後）
"""
import io, json, os, re, sys, time, uuid
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from playwright.sync_api import sync_playwright, TimeoutError as PwTimeout
import openpyxl
from openpyxl.styles import PatternFill

# ── 路徑 ─────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(__file__)
EXCEL_PATH  = os.path.join(SCRIPT_DIR, "..", "data", "cat_foods_database.xlsx")
RAW_DIR     = os.path.join(SCRIPT_DIR, "data", "raw")
os.makedirs(RAW_DIR, exist_ok=True)

# ── 爬蟲設定 ──────────────────────────────────────────────────
BASE_URL    = "https://shop.petpark.com.tw/dryfood/hot-cat-dryfood"
DELAY       = 1.8   # 秒/頁

# ── 評分相關（同 03_scorer.py）────────────────────────────────
def get_label(total):
    if total >= 80: return "優質主食"
    if total >= 65: return "不錯的選擇"
    if total >= 50: return "可以接受"
    return "需謹慎"

def calc_dm(pct, moisture):
    """乾物比換算"""
    if pct is None or moisture is None: return None
    if moisture >= 100: return None
    return round(pct / (1 - moisture / 100), 1)

def calc_carb_dm(protein_dm, fat_dm, ash_dm, fiber_dm):
    vals = [protein_dm, fat_dm, ash_dm]
    if any(v is None for v in vals): return None
    fiber = fiber_dm or 0
    carb = 100 - protein_dm - fat_dm - ash_dm - fiber
    return round(max(0, carb), 1)

def score_food(food):
    p = food.get("protein_dm_pct") or 0
    c = food.get("carb_dm_pct") or 100
    f = food.get("fat_dm_pct") or 0
    a = food.get("ash_pct") or 0
    has_grain        = food.get("has_grain", False)
    is_aafco         = food.get("is_aafco_certified", False)
    has_ingredients  = bool(food.get("ingredients_raw"))
    has_ash_data     = food.get("ash_pct") is not None
    has_fat_data     = food.get("fat_dm_pct") is not None

    total = 0
    bd = {}

    # 蛋白質 30
    if   p >= 50: pts = 30
    elif p >= 45: pts = 26
    elif p >= 40: pts = 22
    elif p >= 35: pts = 20
    elif p >= 30: pts = 13
    elif p >= 26: pts = 7
    else:         pts = 2
    total += pts; bd["protein"] = pts

    # 碳水 15
    if   c <= 10: pts = 15
    elif c <= 20: pts = 12
    elif c <= 30: pts = 9
    elif c <= 40: pts = 8
    else:         pts = 2
    total += pts; bd["carb"] = pts

    # 脂肪 10
    if has_fat_data and f > 0:
        if   20 <= f <= 40: pts = 10
        elif 15 <= f < 20 or 40 < f <= 50: pts = 7
        elif  9 <= f < 15:  pts = 4
        else:                pts = 1
    else: pts = 0
    total += pts; bd["fat"] = pts

    # 透明度 30
    t = (18 if is_aafco else 0) + (7 if has_ingredients else 0) + (5 if has_ash_data else 0)
    total += t; bd["transparency"] = t

    # 無穀 8
    gpts = 0 if has_grain else 8
    total += gpts; bd["no_grain"] = gpts

    # 灰分品質 7
    if has_ash_data:
        if   a <= 8:  pts = 7
        elif a <= 10: pts = 3
        else:         pts = 0
    else: pts = 0
    total += pts; bd["ash"] = pts

    total = max(0, min(100, total))
    food["score_total"]     = total
    food["score_label"]     = get_label(total)
    food["score_protein"]   = bd["protein"]
    food["score_carb"]      = bd["carb"]
    food["score_fat"]       = bd["fat"]
    food["score_transparency"] = bd["transparency"]
    food["score_grain"]     = bd["no_grain"]
    food["score_ash_quality"] = bd["ash"]
    return food


# ── 解析成分 ──────────────────────────────────────────────────
def parse_float_pct(text, keyword):
    """從整頁文字找 keyword 後面的第一個數字"""
    patterns = [
        rf"{keyword}[^0-9]*?(\d+\.?\d*)\s*[％%]",
        rf"{keyword}\s*[(（][^)]*[)）]\s*(\d+\.?\d*)[%％]",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            return float(m.group(1))
    return None

GRAIN_KW = ["玉米", "小麥", "大麥", "燕麥", "白米", "糙米", "高粱",
            "corn", "wheat", "barley", "oat", "rice", "sorghum"]

LIFE_STAGE_MAP = {
    "kitten": ["幼貓", "kitten", "幼齡", "離乳", "K36", "BC34"],
    "senior": ["熟齡", "老貓", "senior", "7+", "高齡", "11+"],
    "all":    ["全齡", "all life", "all age", "各年齡"],
}

def guess_life_stage(text):
    t = text.lower()
    for stage, kws in LIFE_STAGE_MAP.items():
        if any(k.lower() in t for k in kws):
            return stage
    return "adult"

def guess_brand(name):
    known = [
        "法國皇家", "Royal Canin", "希爾思", "Hills", "冠能", "PRO PLAN", "Purina",
        "耐吉斯", "Nutrience", "自然平衡", "奇境", "Zignature", "魏大夫",
        "藍饌", "Blue Buffalo", "愛肯拿", "ACANA", "歐睿健", "Orijen", "ORIJEN",
        "Farmina", "法米納", "Wellness", "威樂", "汪喵星球", "御天貓", "原點",
        "無敵", "怡親", "K9 Natural", "K9Natural", "Petlife", "晶饌",
        "鮮樂嚐", "惜時", "seeds", "Seeds", "addiction", "Addiction", "自然癮食",
        "滋益巔峰", "Ziwi", "ZIWI", "Go!", "GO!", "耐吉斯", "Nutrience",
        "Instinct", "本能", "史黛拉", "Stella", "Merrick", "瑪瑞克",
        "Canidae", "卡比", "Solid Gold", "固德", "Nulo", "紐樂",
        "Fromm", "法蘭", "Nutro", "美士", "Eukanuba", "優卡",
        "NOW Fresh", "原鮮", "iams", "Iams", "寵愛一生", "WDJ",
        "耐吉斯", "皇家", "Applaws", "台灣", "Pronature",
    ]
    for b in known:
        if b.lower() in name.lower():
            return b
    # 取第一段（空白或括號前）
    m = re.match(r"^([^\s（(【\[]+)", name)
    return m.group(1) if m else name[:5]


# ── Step 1：讀現有 Excel，建名稱集合 ─────────────────────────
def load_existing_names(excel_path):
    if not os.path.exists(excel_path):
        return set(), set()
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    name_col = headers.index("飼料名稱") + 1 if "飼料名稱" in headers else 2
    url_col  = headers.index("來源網址") + 1 if "來源網址" in headers else None

    names, urls = set(), set()
    for row in range(2, ws.max_row + 1):
        n = ws.cell(row, name_col).value
        if n:
            names.add(str(n).strip())
        if url_col:
            u = ws.cell(row, url_col).value
            if u:
                urls.add(str(u).strip())
    print(f"現有資料庫：{len(names)} 筆（比對用）")
    return names, urls


def is_duplicate(name, url, existing_names, existing_urls):
    if url in existing_urls:
        return True
    # 名稱前 15 字相同就算重複（不同容量的同款）
    prefix = re.sub(r"\s*\d+[\s.]?\d*\s*(公斤|kg|磅|lb|克|g|ml|L)\b.*", "", name, flags=re.I).strip()
    for en in existing_names:
        ep = re.sub(r"\s*\d+[\s.]?\d*\s*(公斤|kg|磅|lb|克|g|ml|L)\b.*", "", en, flags=re.I).strip()
        if prefix and ep and (prefix == ep or prefix[:12] == ep[:12]):
            return True
    return False


# ── Step 2：爬商品清單（所有分頁）────────────────────────────
def get_all_product_links(page):
    all_items = []  # [(name, url), ...]
    p = 1
    while True:
        url = BASE_URL if p == 1 else f"{BASE_URL}?p={p}"
        print(f"  分頁 {p}：{url}")
        try:
            page.goto(url, wait_until="networkidle", timeout=25000)
            page.wait_for_timeout(2500)
        except PwTimeout:
            print(f"  ⚠️  分頁 {p} 逾時，停止")
            break

        items = page.evaluate("""() => {
            const els = document.querySelectorAll('.product-item-name a, .product-name a, h2 a, .item-title a, [class*=product-item] a[href]');
            return [...els].map(e => ({ name: e.textContent.trim(), href: e.href }))
                           .filter(x => x.name && x.href.includes('shop.petpark'));
        }""")

        if not items:
            print(f"  分頁 {p} 無商品，結束")
            break

        all_items.extend(items)
        print(f"    取得 {len(items)} 筆，累計 {len(all_items)} 筆")

        # 有沒有下一頁
        has_next = page.evaluate(f"""() => !!document.querySelector('a[href*="?p={p+1}"]')""")
        if not has_next:
            break
        p += 1
        time.sleep(DELAY)

    return all_items


# ── Step 3：爬單一商品頁 ──────────────────────────────────────
def scrape_product(page, name, url):
    try:
        page.goto(url, wait_until="networkidle", timeout=25000)
        page.wait_for_timeout(2000)
    except PwTimeout:
        return None

    text = page.inner_text("body")

    protein_pct  = parse_float_pct(text, "粗蛋白質?")
    fat_pct      = parse_float_pct(text, "粗脂肪")
    fiber_pct    = parse_float_pct(text, "粗纖維")
    moisture_pct = parse_float_pct(text, "水分")
    ash_pct      = parse_float_pct(text, "粗?灰分")

    if not any([protein_pct, fat_pct, moisture_pct]):
        return None

    # 成分表
    ing_m = re.search(r"(?:成分|原料)[：:＊*]?\s*(.{20,600}?)(?:\n\n|保證|分析|商品規格|$)", text, re.DOTALL)
    ingredients_raw = re.sub(r"\s+", " ", ing_m.group(1).strip())[:500] if ing_m else None

    is_aafco = bool(re.search(r"AAFCO|aafco", text))
    has_grain_val = any(k in (ingredients_raw or "").lower() for k in GRAIN_KW)

    # 乾物比
    dm_moisture = moisture_pct
    protein_dm = calc_dm(protein_pct, dm_moisture)
    fat_dm     = calc_dm(fat_pct,     dm_moisture)
    fiber_dm   = calc_dm(fiber_pct,   dm_moisture)
    ash_dm     = calc_dm(ash_pct,     dm_moisture)
    carb_dm    = calc_carb_dm(protein_dm, fat_dm, ash_dm, fiber_dm)

    food = {
        "id":               str(uuid.uuid4()),
        "name":             name,
        "brand":            guess_brand(name),
        "life_stage":       guess_life_stage(name + " " + text[:300]),
        "food_type":        "dry",
        "source_url":       url,
        "ingredients_raw":  ingredients_raw,
        "protein_pct":      protein_pct,
        "fat_pct":          fat_pct,
        "fiber_pct":        fiber_pct,
        "moisture_pct":     moisture_pct,
        "ash_pct":          ash_pct,
        "protein_dm_pct":   protein_dm,
        "fat_dm_pct":       fat_dm,
        "carb_dm_pct":      carb_dm,
        "ash_dm_pct":       ash_dm,
        "has_grain":        has_grain_val,
        "is_aafco_certified": is_aafco,
        "has_ingredient_list": bool(ingredients_raw),
        "has_ash_listed":   ash_pct is not None,
    }
    return score_food(food)


# ── Step 4：把新資料追加到 Excel ──────────────────────────────
COLUMNS = [
    ("brand",               "品牌"),
    ("name",                "飼料名稱"),
    ("life_stage",          "適用階段"),
    ("score_total",         "總分"),
    ("score_label",         "評級"),
    ("protein_dm_pct",      "蛋白質% (乾重)"),
    ("fat_dm_pct",          "脂肪% (乾重)"),
    ("carb_dm_pct",         "碳水% (乾重)"),
    ("ash_dm_pct",          "灰分% (乾重)"),
    ("moisture_pct",        "水分%"),
    ("protein_pct",         "蛋白質% (標示)"),
    ("fat_pct",             "脂肪% (標示)"),
    ("fiber_pct",           "纖維% (標示)"),
    ("ash_pct",             "灰分% (標示)"),
    ("has_grain",           "含穀物"),
    ("is_aafco_certified",  "AAFCO 認證"),
    ("has_ingredient_list", "有成分表"),
    ("has_ash_listed",      "有標示灰分"),
    ("score_protein",       "蛋白質分"),
    ("score_carb",          "碳水分"),
    ("score_fat",           "脂肪分"),
    ("score_transparency",  "透明度分"),
    ("score_grain",         "無穀分"),
    ("score_ash_quality",   "灰分品質分"),
    ("source_url",          "來源網址"),
    ("id",                  "UUID"),
]

STAGE_LABEL = {"kitten": "幼貓", "adult": "成貓", "senior": "熟齡貓", "all": "全齡"}

NEW_ROW_FILL = PatternFill("solid", fgColor="D9EAD3")  # 淡綠色標示新增列

def append_to_excel(excel_path, new_foods):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    for food in new_foods:
        row = ws.max_row + 1
        food["life_stage"] = STAGE_LABEL.get(food.get("life_stage", "adult"), food.get("life_stage", ""))
        for key, label in COLUMNS:
            col = headers.index(label) + 1 if label in headers else None
            if col is None:
                continue
            val = food.get(key)
            if isinstance(val, bool):
                val = "是" if val else "否"
            c = ws.cell(row=row, column=col, value=val)
            c.fill = NEW_ROW_FILL

    wb.save(excel_path)
    print(f"✅ 已追加 {len(new_foods)} 筆到 {excel_path}")


# ── 主流程 ────────────────────────────────────────────────────
def main():
    existing_names, existing_urls = load_existing_names(EXCEL_PATH)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="zh-TW",
        )
        list_page   = ctx.new_page()
        detail_page = ctx.new_page()

        # Step A：抓所有商品清單
        print("\n── Step 1：抓商品清單 ──────────────────────────────────")
        all_items = get_all_product_links(list_page)
        print(f"\n總計找到 {len(all_items)} 個商品")

        # Step B：過濾新品
        new_items = []
        for item in all_items:
            if not is_duplicate(item["name"], item["href"], existing_names, existing_urls):
                new_items.append(item)
        print(f"扣除重複後，新品 {len(new_items)} 個\n")

        # Step C：爬各新品詳細頁
        print("── Step 2：爬新品成分 ──────────────────────────────────")
        results = []
        no_data = []

        for i, item in enumerate(new_items, 1):
            name = item["name"]
            url  = item["href"]
            print(f"[{i}/{len(new_items)}] {name[:45]}")
            food = scrape_product(detail_page, name, url)
            if food:
                results.append(food)
                print(f"  ✅ 蛋白質(標示)={food.get('protein_pct')}%  水分={food.get('moisture_pct')}%  → 評分 {food['score_total']} [{food['score_label']}]")
            else:
                no_data.append(name)
                print(f"  ⚠️  無法解析成分，略過")
            time.sleep(DELAY)

        browser.close()

    # Step D：存到 Excel
    print(f"\n── Step 3：輸出到 Excel ────────────────────────────────")
    print(f"成功解析：{len(results)} 筆")
    print(f"無成分資料：{len(no_data)} 筆")

    if results:
        # 同時存一份 raw JSON 備份
        raw_path = os.path.join(RAW_DIR, f"petpark_new_{time.strftime('%Y%m%d_%H%M')}.json")
        with open(raw_path, "w", encoding="utf-8", newline="\n") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print(f"JSON 備份：{raw_path}")

        append_to_excel(EXCEL_PATH, results)

    if no_data:
        print(f"\n無成分資料的商品（{len(no_data)} 筆）：")
        for n in no_data[:20]:
            print(f"  - {n[:60]}")

    print(f"\n完成！請開啟 Excel 確認後，告訴我要上傳哪些筆到 Supabase。")


if __name__ == "__main__":
    main()
