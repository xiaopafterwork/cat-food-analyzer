# -*- coding: utf-8 -*-
"""
clean_and_rebuild.py
1. 去除品名中的重量/規格/貓飼料等無關資訊
2. 去重（綠底與白底之間、以及全部資料內部）
3. 重新計算評分
4. 用規則式邏輯生成內容評價（編輯 Agent）
5. 輸出乾淨的 Excel
"""
import io, re, sys, uuid
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

EXCEL_PATH = r"C:\Users\P\cat-food-analyzer\data\cat_foods_database.xlsx"
OUT_PATH   = r"C:\Users\P\cat-food-analyzer\data\cat_foods_database.xlsx"  # overwrite

# ── 欄位順序 ─────────────────────────────────────────────────
COLS = [
    "品牌", "飼料名稱", "適用階段", "總分", "評級",
    "蛋白質% (乾重)", "脂肪% (乾重)", "碳水% (乾重)", "灰分% (乾重)", "水分%",
    "蛋白質% (標示)", "脂肪% (標示)", "纖維% (標示)", "灰分% (標示)",
    "含穀物", "AAFCO 認證", "有成分表", "有標示灰分",
    "蛋白質分", "碳水分", "脂肪分", "透明度分", "無穀分", "灰分品質分",
    "內容評價",   # ← 新增欄位
    "來源網址", "UUID",
]

COL_IDX = {c: i+1 for i, c in enumerate(COLS)}

# ── 品名清洗 ─────────────────────────────────────────────────
NOISE_PATTERNS = [
    r"\s*\(貓飼料\)",
    r"\s*（貓飼料）",
    r"\s*\(貓糧\)",
    r"\s*\d+\.?\d*\s*(公斤|公克|kg|KG|Kg|磅|lb|LB|克|g|G|ml|ML|L|入|包|袋)\b.*",
    r"\s+\d+\s*$",  # 尾端殘留數字
]

def clean_name(name: str) -> str:
    if not name:
        return name
    for pat in NOISE_PATTERNS:
        name = re.sub(pat, "", name, flags=re.IGNORECASE)
    return name.strip()


# ── 去重 key ─────────────────────────────────────────────────
def dedup_key(name: str) -> str:
    """取清洗後品名前 20 字作為去重 key（忽略空白大小寫）"""
    n = re.sub(r"\s+", "", name.lower())
    return n[:20]


# ── 評分公式（同 03_scorer.py）────────────────────────────────
def score_food(row: dict) -> dict:
    p = row.get("蛋白質% (乾重)") or 0
    c = row.get("碳水% (乾重)") or 100
    f = row.get("脂肪% (乾重)") or 0
    a = row.get("灰分% (乾重)") or 0
    has_grain  = row.get("含穀物") in ("是", True)
    is_aafco   = row.get("AAFCO 認證") in ("是", True)
    has_ingred = row.get("有成分表") in ("是", True)
    has_ash    = row.get("有標示灰分") in ("是", True)
    has_fat    = row.get("脂肪% (乾重)") is not None

    # 蛋白質 30
    if   p >= 50: pp = 30
    elif p >= 45: pp = 26
    elif p >= 40: pp = 22
    elif p >= 35: pp = 20
    elif p >= 30: pp = 13
    elif p >= 26: pp = 7
    else:         pp = 2

    # 碳水 15
    if   c <= 10: cp = 15
    elif c <= 20: cp = 12
    elif c <= 30: cp = 9
    elif c <= 40: cp = 8
    else:         cp = 2

    # 脂肪 10
    if has_fat and f > 0:
        if   20 <= f <= 40: fp = 10
        elif 15 <= f < 20 or 40 < f <= 50: fp = 7
        elif  9 <= f < 15: fp = 4
        else: fp = 1
    else:
        fp = 0

    # 透明度 30
    tp = (18 if is_aafco else 0) + (7 if has_ingred else 0) + (5 if has_ash else 0)

    # 無穀 8
    gp = 0 if has_grain else 8

    # 灰分品質 7
    if has_ash:
        if   a <= 8:  ap = 7
        elif a <= 10: ap = 3
        else:         ap = 0
    else:
        ap = 0

    total = max(0, min(100, pp + cp + fp + tp + gp + ap))

    if   total >= 80: label = "優質主食"
    elif total >= 65: label = "不錯的選擇"
    elif total >= 50: label = "可以接受"
    else:             label = "需謹慎"

    row["總分"]    = total
    row["評級"]    = label
    row["蛋白質分"] = pp
    row["碳水分"]  = cp
    row["脂肪分"]  = fp
    row["透明度分"] = tp
    row["無穀分"]  = gp
    row["灰分品質分"] = ap
    return row


# ── 內容評價生成（規則式，編輯 Agent）────────────────────────
# 蛋白質評語（7種措辭避免重複）
def _protein_line(p, dm=True):
    src = "乾重換算後" if dm else "標示"
    if p is None or p == 0:
        return "蛋白質資料未標示，建議向廠商確認實際含量。"
    if p >= 55:
        return f"蛋白質高達 {p:.0f}%（{src}），超越大多數市售主食，非常適合肉食需求較高的貓咪。"
    if p >= 50:
        return f"蛋白質 {p:.0f}%（{src}），含量優秀，貓咪攝取主要能量來源充足。"
    if p >= 45:
        return f"蛋白質 {p:.0f}%（{src}），高於平均水準，是優質動物蛋白的良好來源。"
    if p >= 40:
        return f"蛋白質 {p:.0f}%（{src}），符合貓咪日常基本需求。"
    if p >= 35:
        return f"蛋白質 {p:.0f}%（{src}），中等水準，長期餵食建議搭配高蛋白副食補充。"
    if p >= 26:
        return f"蛋白質 {p:.0f}%（{src}），略低於理想範圍，建議作為副食而非唯一主食。"
    return f"蛋白質僅 {p:.0f}%（{src}），含量偏低，不建議長期作為主食單獨餵食。"


def _carb_line(c):
    if c is None:
        return "碳水化合物因成分資料不完整，無法精確計算。"
    if c <= 5:
        return f"碳水化合物極低（約 {c:.0f}%），完全符合貓咪肉食性生理設計。"
    if c <= 10:
        return f"碳水化合物僅 {c:.0f}%，屬低碳配方，對貓咪代謝友善。"
    if c <= 20:
        return f"碳水化合物 {c:.0f}%，屬中等水準，健康成貓偶爾食用無虞。"
    if c <= 30:
        return f"碳水化合物 {c:.0f}%，偏高，建議避免長期單一餵食，以免胰島素負擔過重。"
    if c <= 40:
        return f"碳水化合物達 {c:.0f}%，明顯偏高，貓咪長期食用可能增加體重及糖尿病風險。"
    return f"碳水化合物高達 {c:.0f}%，遠超貓咪所需，屬穀物/澱粉比例過高的配方，需謹慎選用。"


def _summary_line(row):
    score = row.get("總分") or 0
    has_grain = row.get("含穀物") in ("是", True)
    is_aafco  = row.get("AAFCO 認證") in ("是", True)
    p = row.get("蛋白質% (乾重)") or 0
    c = row.get("碳水% (乾重)") or 100

    if score >= 80:
        base = "整體評分優良，推薦作為日常主食。"
    elif score >= 65:
        base = "整體表現不錯，是值得考慮的日常選擇。"
    elif score >= 50:
        base = "整體尚可接受，適合作為搭配副食使用。"
    else:
        base = "整體評分偏低，建議謹慎評估是否長期餵食。"

    notes = []
    if is_aafco:
        notes.append("通過 AAFCO 國際營養標準")
    if not has_grain:
        notes.append("無穀物配方")
    if p >= 45:
        notes.append("高蛋白來源")
    if c > 30:
        notes.append("⚠️ 碳水偏高需注意")

    if notes:
        return base + "（" + "、".join(notes) + "）"
    return base


def generate_review(row: dict) -> str:
    p = row.get("蛋白質% (乾重)") or row.get("蛋白質% (標示)")
    use_dm = row.get("蛋白質% (乾重)") is not None
    c = row.get("碳水% (乾重)")
    line1 = _protein_line(p, dm=use_dm)
    line2 = _carb_line(c)
    line3 = _summary_line(row)
    return f"{line1}\n{line2}\n{line3}"


# ── 主流程 ───────────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # 讀取原始欄位
    orig_headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    oh = {v: i for i, v in enumerate(orig_headers)}  # header → col index (0-based)

    def cell_val(row_vals, key):
        idx = oh.get(key)
        return row_vals[idx] if idx is not None and idx < len(row_vals) else None

    # 讀取所有資料列 + 記錄是否綠底
    all_rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(vals):  # 空列跳過
            continue
        is_green = (ws.cell(r, 1).fill.fgColor.rgb == "FFD9EAD3")
        row_dict = {k: cell_val(vals, k) for k in orig_headers if k}
        row_dict["_is_green"] = is_green
        all_rows.append(row_dict)

    print(f"讀取 {len(all_rows)} 筆（白底 {sum(1 for r in all_rows if not r['_is_green'])}，綠底 {sum(1 for r in all_rows if r['_is_green'])}）")

    # Step 1: 清洗品名（所有列）
    for row in all_rows:
        row["飼料名稱"] = clean_name(row.get("飼料名稱") or "")

    # Step 1b: 自動補算乾重（當乾重為空但標示值+水分有值時）
    def calc_dm(pct, moisture):
        if pct is None or moisture is None or moisture >= 100:
            return None
        return round(pct / (1 - moisture / 100), 1)

    for row in all_rows:
        moisture = row.get("水分%")
        # 蛋白質乾重
        if row.get("蛋白質% (乾重)") is None and row.get("蛋白質% (標示)") and moisture:
            row["蛋白質% (乾重)"] = calc_dm(row["蛋白質% (標示)"], moisture)
        # 脂肪乾重
        if row.get("脂肪% (乾重)") is None and row.get("脂肪% (標示)") and moisture:
            row["脂肪% (乾重)"] = calc_dm(row["脂肪% (標示)"], moisture)
        # 碳水（需要蛋白/脂肪/灰分乾重）
        if row.get("碳水% (乾重)") is None:
            p_dm = row.get("蛋白質% (乾重)")
            f_dm = row.get("脂肪% (乾重)")
            a_dm = calc_dm(row.get("灰分% (標示)"), moisture)
            fi_dm = calc_dm(row.get("纖維% (標示)"), moisture)
            if all(v is not None for v in [p_dm, f_dm, a_dm]):
                row["碳水% (乾重)"] = round(max(0, 100 - p_dm - f_dm - a_dm - (fi_dm or 0)), 1)

    # Step 2: 去重（保留第一次出現，優先保留白底）
    # 白底先，綠底後
    white_rows = [r for r in all_rows if not r["_is_green"]]
    green_rows = [r for r in all_rows if r["_is_green"]]

    seen_keys = set()
    seen_urls = set()
    kept = []

    for row in white_rows + green_rows:
        name = row.get("飼料名稱") or ""
        url  = row.get("來源網址") or ""
        key  = dedup_key(name)
        is_dup = False

        if url and url in seen_urls:
            is_dup = True
        elif key and key in seen_keys:
            is_dup = True

        if not is_dup:
            seen_keys.add(key)
            if url:
                seen_urls.add(url)
            kept.append(row)

    removed = len(all_rows) - len(kept)
    print(f"去重移除 {removed} 筆，剩餘 {len(kept)} 筆")

    # Step 3: 重新計算評分
    for row in kept:
        score_food(row)

    # Step 4: 生成內容評價
    for row in kept:
        row["內容評價"] = generate_review(row)

    # Step 5: 補齊 UUID（若缺）
    for row in kept:
        if not row.get("UUID"):
            row["UUID"] = str(uuid.uuid4())

    # Step 6: 輸出新 Excel
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "貓飼料資料庫"

    # 表頭樣式
    header_fill = PatternFill("solid", fgColor="2D4A2E")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for ci, col in enumerate(COLS, 1):
        c = new_ws.cell(1, ci, col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    new_ws.row_dimensions[1].height = 30

    # 分數色碼
    def score_fill(score):
        if score is None: return None
        if score >= 80: return PatternFill("solid", fgColor="C8E6C9")
        if score >= 65: return PatternFill("solid", fgColor="FFF9C4")
        if score >= 50: return PatternFill("solid", fgColor="FFE0B2")
        return PatternFill("solid", fgColor="FFCDD2")

    # 寫入資料
    orig_white  = set(id(r) for r in white_rows)
    green_fill  = PatternFill("solid", fgColor="D9EAD3")

    for ri, row in enumerate(kept, 2):
        is_new = row.get("_is_green", False)
        s = row.get("總分")

        for ci, col in enumerate(COLS, 1):
            val = row.get(col)
            # bool 轉中文
            if isinstance(val, bool):
                val = "是" if val else "否"
            c = new_ws.cell(ri, ci, val)
            c.alignment = Alignment(vertical="center", wrap_text=(col == "內容評價"))
            # 底色
            if col == "總分" and s is not None:
                c.fill = score_fill(s)
            elif is_new:
                c.fill = green_fill
            # 評價欄字體小一號
            if col == "內容評價":
                c.font = Font(size=9)

        if col == "內容評價":
            new_ws.row_dimensions[ri].height = 55

    # 欄寬
    col_widths = {
        "品牌": 14, "飼料名稱": 38, "適用階段": 8, "總分": 7, "評級": 10,
        "蛋白質% (乾重)": 12, "脂肪% (乾重)": 12, "碳水% (乾重)": 12, "灰分% (乾重)": 12, "水分%": 8,
        "蛋白質% (標示)": 12, "脂肪% (標示)": 12, "纖維% (標示)": 12, "灰分% (標示)": 12,
        "含穀物": 8, "AAFCO 認證": 10, "有成分表": 8, "有標示灰分": 10,
        "蛋白質分": 8, "碳水分": 8, "脂肪分": 8, "透明度分": 8, "無穀分": 8, "灰分品質分": 10,
        "內容評價": 50,
        "來源網址": 40, "UUID": 36,
    }
    for ci, col in enumerate(COLS, 1):
        new_ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 12)

    new_ws.freeze_panes = "A2"

    new_wb.save(OUT_PATH)
    print(f"\n✅ 完成！共 {len(kept)} 筆資料")
    print(f"   白底（原有）: {sum(1 for r in kept if not r.get('_is_green'))} 筆")
    print(f"   綠底（新增）: {sum(1 for r in kept if r.get('_is_green'))} 筆")
    print(f"   輸出：{OUT_PATH}")

    # 分數分布報告
    labels = {}
    for r in kept:
        lbl = r.get("評級") or "未知"
        labels[lbl] = labels.get(lbl, 0) + 1
    print("\n評分分布：")
    for lbl in ["優質主食", "不錯的選擇", "可以接受", "需謹慎", "未知"]:
        if lbl in labels:
            print(f"  {lbl}: {labels[lbl]} 筆")


if __name__ == "__main__":
    main()
