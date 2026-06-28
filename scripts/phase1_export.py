# -*- coding: utf-8 -*-
"""
phase1_export.py — 自動找 URL，產出待審核 Excel

流程：
  1. 讀取 targets.json 現有清單
  2. 每筆自動查：官網URL（品牌對照表 or DuckDuckGo）+ momo 第二來源
  3. 輸出 Excel 供你審核
  4. 你在 Excel：確認URL / 刪不要的行 / 手動新增新品 / 批准欄填 Y
  5. 存檔後執行 python phase2_crawl.py

用法：
  python phase1_export.py           # 自動找 URL
  python phase1_export.py --skip-url  # 跳過URL搜尋（快速，URL留空手填）
"""

import io, json, os, sys, argparse
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

sys.path.insert(0, os.path.dirname(__file__))
from crawlers.url_finder import find_urls

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("請先安裝：pip install openpyxl"); sys.exit(1)

SCRIPT_DIR   = os.path.dirname(__file__)
TARGETS_FILE = os.path.join(SCRIPT_DIR, "targets.json")
OUT_DIR      = os.path.join(SCRIPT_DIR, "data/review")
os.makedirs(OUT_DIR, exist_ok=True)

HEADERS_COLS = [
    "品牌", "飼料名稱",
    "官網URL（優先）", "第二來源URL（PChome）",
    "批准\n(Y=批准 N=跳過)", "備註",
]
COL_WIDTHS = [22, 30, 45, 45, 14, 25]
MANUAL_ROWS = 30

# 顏色
GREEN  = "3D5A3E"; LGREEN = "E8F5E9"; LGRAY = "F5F5F5"
YELLOW = "FFF9C4"; WHITE  = "FFFFFF"; WARN  = "FFF3CD"

def border():
    s = Side(border_style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def hcell(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=True, color=WHITE, size=10)
    c.fill      = PatternFill("solid", fgColor=GREEN)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border    = border()

def section(ws, row, label, color=GREEN):
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(bold=True, color=WHITE, size=11)
    c.fill      = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS_COLS))
    ws.row_dimensions[row].height = 22

def drow(ws, row, vals, bg=WHITE, url_warn=False):
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        fill_color = WARN if (url_warn and col in (3, 4)) else bg
        c.fill      = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(vertical="center", wrap_text=(col == 6))
        c.border    = border()
        if col in (3, 4) and val and str(val).startswith("http"):
            c.font = Font(color="1155CC", underline="single", size=9)
        elif col in (3, 4) and not val:
            c.font = Font(color="CC0000", italic=True, size=9)


def build_excel(targets: list[dict]) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "待審核清單"
    ws.freeze_panes = "A4"
    row = 1

    # 說明列
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS_COLS))
    c = ws.cell(row=row, column=1,
        value="📋 操作說明：① 確認URL正確（紅字=找不到，請手填）"
              "  ② 刪掉不需要的行  ③ 在「手動新增」區加新品"
              "  ④ 批准欄填 Y  ⑤ 存檔 → python phase2_crawl.py")
    c.font = Font(italic=True, color="555555", size=9)
    c.fill = PatternFill("solid", fgColor="FFFDE7")
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1

    # ── 爬蟲清單區 ──
    section(ws, row, f"▶ 爬蟲清單（{len(targets)} 筆，來自 targets.json）")
    row += 1
    for col, h in enumerate(HEADERS_COLS, 1):
        hcell(ws, row, col, h)
    ws.row_dimensions[row].height = 30
    row += 1

    for i, t in enumerate(targets):
        o_url = t.get("official_url") or ""
        s_url = t.get("second_url") or t.get("pchome_url") or t.get("momo_url") or ""
        url_warn = not o_url  # 官網找不到就標橘
        drow(ws, row, [
            t.get("brand", ""), t.get("name", ""),
            o_url, s_url, "", "",
        ], bg=WHITE if i % 2 == 0 else LGRAY, url_warn=url_warn)
        row += 1

    row += 1  # 空行

    # ── 手動新增區 ──
    section(ws, row, f"▶ 手動新增（最多 {MANUAL_ROWS} 筆，填完批准欄填 Y）", color="2E6B4E")
    row += 1
    for col, h in enumerate(HEADERS_COLS, 1):
        hcell(ws, row, col, h)
    ws.row_dimensions[row].height = 30
    row += 1

    for i in range(MANUAL_ROWS):
        bg = YELLOW if i % 2 == 0 else "FFFEF0"
        drow(ws, row, ["", "", "", "", "", ""], bg=bg)
        row += 1

    for col, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    date_str = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = os.path.join(OUT_DIR, f"待審核_{date_str}.xlsx")
    wb.save(out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--skip-url", action="store_true", help="跳過URL自動搜尋")
    args = parser.parse_args()

    targets = []
    if os.path.exists(TARGETS_FILE):
        targets = json.load(open(TARGETS_FILE, encoding="utf-8"))
        print(f"讀取 targets.json：{len(targets)} 筆")
    else:
        print("找不到 targets.json，將產出空白清單")

    if not args.skip_url and targets:
        print(f"\n自動搜尋 URL（每筆約 3 秒）…\n{'─'*50}")
        for i, t in enumerate(targets):
            brand = t.get("brand", "")
            name  = t.get("name", "")
            print(f"  [{i+1}/{len(targets)}] {brand} — {name}")

            # 已有 URL 就跳過
            if t.get("official_url") and (t.get("second_url") or t.get("momo_url")):
                print(f"    已有URL，跳過搜尋")
                continue

            urls = find_urls(brand, name)
            if not t.get("official_url"):
                t["official_url"] = urls["official_url"]
            if not t.get("second_url") and not t.get("momo_url"):
                t["second_url"] = urls["second_url"]

            o = t.get("official_url") or "❌ 找不到"
            s = t.get("second_url")   or "—"
            print(f"    官網: {str(o)[:65]}")
            print(f"    pchome: {str(s)[:65]}")
    else:
        print("\n（跳過URL搜尋，請在 Excel 手動填入）")

    out = build_excel(targets)
    print(f"\n{'─'*50}")
    print(f"✅ Excel 已輸出：{out}")
    print("\n接下來：")
    print("  1. 開啟 Excel，確認 URL（紅字=找不到，需手填）")
    print("  2. 刪掉不需要的行，在黃色區新增品項")
    print("  3. 批准欄填 Y → 存檔")
    print("  4. python phase2_crawl.py")


if __name__ == "__main__":
    main()
